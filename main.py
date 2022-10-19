from weather import get_weather
from cinemas import get_cinemas
from museum import get_museums
from movieIMDb import get_movies
from twitter import get_tweets
from Selenium_TVGuia import Selenium_bot
from DataBase import GetPandasDataFrame
from DataBase import DfToSql
import xlsxwriter
import openpyxl as xl


def Excel_Movies(movies,workbook):
    W_Movies = workbook.add_worksheet("movies")
    W_Movies.write(0, 0, "title") #Title
    W_Movies.write(0, 1, "release_date") #Release date
    W_Movies.write(0, 2, "rating") #Rating
    W_Movies.write(0, 3, "duration") #Duration
    W_Movies.write(0, 4, "poster") #Poster
    W_Movies.write(0, 5, "genre") #Genre
    W_Movies.write(0, 6, "summary") #Summary
    W_Movies.write(0, 7, "score") #Score
    W_Movies.write(0, 8, "director") #Director
    W_Movies.write(0, 9, "cast") #Cast
    W_Movies.write(0, 10, "trailer") #Trailer
    row = 1
    for movie in movies:
        actors = ""

        #Fill the first 8 columns
        for i in range(0,9):
            W_Movies.write(row, i, movie[i])
        
        try:
            for actor in movie[9]:
                actors += " " + actor
        except:
            actors = ""
        W_Movies.write(row, 9, actors) #Cast
        W_Movies.write(row, 10, movie[10]) #Trailer
        row += 1
    return


def Excel_Tvguia(TvGuiaInfo,workbook):
    W_Tvguia = workbook.add_worksheet("tvguia")
    #Set initial columns
    W_Tvguia.write(0, 0, "date") #Title
    W_Tvguia.write(0, 1, "timing") #Release date
    W_Tvguia.write(0, 2, "channel") #Rating
    W_Tvguia.write(0, 3, "program_name") #Duration
    row = 1
    #extract Days 
    for a in TvGuiaInfo:
        day = a[0]
        # a=[day n , [    [  [ channel , [[program1, time] [program2, time] [program3, time] ... ]]    [] [] ...(121 channels) ]    []    []   ... (8 time periods)  ]     ]    
        # each time period arrat
        for b in a[1]:
            #each channel in this time period
            for c in b:
                channel = c[0]
                #each program array in this channel
                for d in c[1]:
                    program_name = d[0]
                    program_time = d[1]
                    W_Tvguia.write(row,0,day)
                    W_Tvguia.write(row,1,program_time)
                    W_Tvguia.write(row,2,channel)
                    W_Tvguia.write(row,3,program_name)
                    row += 1
    return


def Excel_Museums(museumsInfo, workbook):
    W_Museums = workbook.add_worksheet("museums")
    W_Museums.write(0, 0, "name") #Museum name
    W_Museums.write(0, 1, "interesting_fact") #Museum best known painting
    W_Museums.write(0, 2, "image") #Museum image
    W_Museums.write(0, 3, "category") #Museum category
    W_Museums.write(0, 4, "description") #Museum description
    W_Museums.write(0, 5, "address") #Museum address
    W_Museums.write(0, 6, "visiting_hours") #Museum hours
    #W_Museums.write(0, 7, "telephone") #Contact information -> We get it from Excel file
    row = 1

    for museum in museumsInfo:
        col = 0
        last = len(museum) - 1
        for info in museum:
            if info != museum[last]:
                W_Museums.write(row, col, info)
                col += 1
        row += 1
    return


def Excel_Weather(weather, workbook):
    W_Weather = workbook.add_worksheet("weather")
    W_Weather.write(0, 0, "location") #Location name
    W_Weather.write(0, 1, "date") #Day
    W_Weather.write(0, 2, "max_min_temperature") #Max/Min temperature
    W_Weather.write(0, 3, "hour") #Hour
    W_Weather.write(0, 4, "temperature") #Temperature (ºC)
    W_Weather.write(0, 5, "image_url") #Image
    W_Weather.write(0, 6, "rain") #Rain (%)
    row = 1

    for location in weather:
        loc = location[0] #Location
        d = location [1] #Date
        temp = location[2] #max_min_temperature
        arr = location[3]
        i = 0
        #hour, temperature, image, rain
        for info in arr:
            col = 3
            for extra in info:
                W_Weather.write(row, 0, loc)
                W_Weather.write(row, 1, d)
                W_Weather.write(row, 2, temp)
                W_Weather.write(row, col, extra)
                col += 1
            row += 1
    return


def Excel_Cinemas(cinemas,workbook):
    W_Cinemas = workbook.add_worksheet("cinemas")

    W_Cinemas.write(0, 0, "date") #Date
    W_Cinemas.write(0, 1, "cinema") #Cinema name
    W_Cinemas.write(0, 2, "films") #Film name
    W_Cinemas.write(0, 3, "time") #Film schedule time

    row = 1

    for cine in cinemas:
        date = cine[0]
        cinema_name = cine[1]
        film_name = cine[2]
        for time in cine[3]:
            W_Cinemas.write(row,0,date)
            W_Cinemas.write(row,1,cinema_name)
            W_Cinemas.write(row,2,film_name)
            W_Cinemas.write(row,3,time)

            row += 1

    return

def Excel_Twitter(TwitterInfo,workbook):
    W_Twitter = workbook.add_worksheet("twitter")
    #Set initial columns
    W_Twitter.write(0, 0, "filmus_name")
    W_Twitter.write(0, 1, "user_id")
    W_Twitter.write(0, 2, "tweet")
    W_Twitter.write(0, 3, "likes")
    W_Twitter.write(0, 4, "retweets")
    row = 1

    for all_tweets in TwitterInfo:
        name = all_tweets[0]
        for tweet in all_tweets[1]:
            # if tweet != "o":
            W_Twitter.write(row, 0, name) #Film/museum name
            W_Twitter.write(row, 1, str(tweet[0])) #User ID
            W_Twitter.write(row, 2, tweet[1]) #Text 
            W_Twitter.write(row, 3, tweet[2]) #Likes
            W_Twitter.write(row, 4, tweet[3]) #Retweets
            row += 1
    return


def Excel_Cultural_Places(workbook):
    # opening the source excel file
    filename ="cultural_places.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
      
    # opening the destination excel file 
    W_Places = workbook.add_worksheet("cultural_places")
      
    # calculate total number of rows and 
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column
      
    # copying the cell values from source 
    # excel file to destination excel file
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row = i, column = j).value
            # print (c)
      
            # writing the read value to destination excel file
            W_Places.write(i-1, j-1, c)

    return


workbook = xlsxwriter.Workbook('Database.xlsx')

print("Obtaining movies info.............")
movies_url = "https://www.imdb.com/showtimes/location?ref_=sh_lc"
movies = get_movies(movies_url)
print("Creating movies Excel.............")
Excel_Movies(movies,workbook)

print("Obtaining museums info.............")
museums_url = "https://es.hoteles.com/go/espana/mejores-museos-de-madrid"
museumsInfo = get_museums(museums_url)
print("Creating museums Excel.............")
Excel_Museums(museumsInfo, workbook)

print("Obtaining TVGuia info.............")
TvGuiaInfo = Selenium_bot()
print("Creating TVGuia Excel.............")
Excel_Tvguia(TvGuiaInfo,workbook)

print("Obtaining Twitter info...........")
tweets_movies = get_tweets(movies)
tweets_museums = get_tweets(museumsInfo)
tweets = tweets_movies + tweets_museums

#[['S', [[1581380547062136832, "@Ironica_ca @sendaabaetxo @MaxTena1 Si. Ir al cine. Quiero ir a ver smile, pero me han dicho que da un miedo que te", 0, 0], [1581380333127155712, '@HormigaXeneize Cada muerto que tuvimos como lateral, Sosa, Fuenzalida, Grana, Smile jara, Buffa etc\n \nNo era muy difícil', 0, 0], [1581380243834621952, '@Evil__Smile Yo quiero que sean novios. Son demasiado bellos los dos ', 0, 0], [1581379983393509376, '@soyeon_smile_ normal peque, es que es&gt;&gt;&gt;', 1, 0], [1581379643957219328, 'quien se viene a ver Smile conmigo ? ', 0, 0], [1581379286568615936, 'Alguien que quiera ir al cine a ver smile ', 0, 0], [1581378659520499713, 'Como que van a ir a ver Smile sin m Esperense a que salga da trabajaaaaaar', 0, 0], [1581378482290163712, 'Quiero ir al cine a ver smile', 2, 1], [1581378196662276096, '@shavafry ¿Es la de Smile? No Sean Toga', 0, 0], [1581377878968893440, '@Evil__Smile No veo a ningún viserys I hay solo al segundo mejor daemon de esa familia.', 0, 0]]], ['B', [[1581379373789499392, 'viendo bullet train x la trama ', 0, 0], [1581377209440563200, '@Perry_McQueen Te recomiendo enormemente Bullet Train, no es terror pero vas a tener morriña de ver Pulp Fiction después.', 0, 0], [1581370988935385088, '@KMubis @PliskeenDR Y fue por un acuerdo que incluyó Bullet Train.', 1, 0], [1581362241416617985, 'Ver Bad Bunny actuar en The Bullet Train son 15min de mi vida que jamás voy a recuperar 😭', 1, 0], [1581350958436847616, '@BytesNsound Bullet train desde dominos', 1, 0], [1581347332695916545, 'Joder la película de bullet train es increíble', 1, 0], [1581344796035100672, 'Bullet Train es una joyita', 1, 0], [1581337933998538754, '137. Bullet Train es tremendamente absurda y precisamente por eso me encanta. Estoy deseando verla otra vez. https://t.co/rtQnu2N1ae', 0, 0], [1581336582254080000, "Vista 'Bullet Train'.\n\nSeguiremos informando. https://t.co/XJoru4QW39 https://t.co/fr8mumIU87", 2, 0], [1581336379753054209, 'Bullet Train. Fantástica comedia negra de acción. Cada uno de sus carismáticos personajes a cual mejor. Hacía tiempo que no disfrutaba tanto de una película. Alocada, divertida y sin complejos. Gran puesta en escena. Me declaro fan absoluto de Limón. Para mí, notable. Un 8. https://t.co/iNvbba76Uk', 1, 0]]]]

print("Creating Twitter Excel.............")
Excel_Twitter(tweets, workbook)

print("Obtaining weather info.............")
weather_url = "https://www.aemet.es/es/eltiempo/prediccion/municipios?p=28&w=t"
weather = get_weather(weather_url)
print("Creating weather Excel.............")
Excel_Weather(weather, workbook)

print("Obtaining cinemas info.............")
cinemas_url = "https://www.imdb.com/showtimes/ES/28912/"
cinemas = get_cinemas(cinemas_url)
print("Creating cinemas Excel.............")
Excel_Cinemas(cinemas,workbook)

print ("Obtaining cultural places info and creating excel.............")
Excel_Cultural_Places(workbook)

workbook.close()

DfToSql(GetPandasDataFrame('Database.xlsx'))